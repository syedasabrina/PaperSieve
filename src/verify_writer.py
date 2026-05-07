# ---------------------------------------------------------------------------
# Verify Writer — xlsx output for verifiability pipeline
# ---------------------------------------------------------------------------
# Responsibilities:
#   1. Create a 4-sheet xlsx if it does not exist:
#        - Definition   : one row per paper, definition prompt fields
#        - Taxonomy     : one row per focal task per paper
#        - Handling     : one row per handling strategy per paper
#        - Summary      : one row per paper, key fields from all three prompts
#   2. Append rows for a new VerifierRecord without rewriting the full file.
#   3. Load already-processed paper IDs for crash recovery.
# ---------------------------------------------------------------------------

from __future__ import annotations
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from src.verify import VerifierRecord


# ---------------------------------------------------------------------------
# Sheet headers
# ---------------------------------------------------------------------------

DEFINITION_HEADERS = [
    "paper_id",
    "subjectivity_def_type",
    "pillar_match",
    "matches_working_definition",
    "definition_gap_identified",
    "subjectivity_distinguished_from",
    "supporting_definition_quote",
    "thought_process",
]

TAXONOMY_HEADERS = [
    "paper_id",
    "task_name",
    "author_task_label",
    "taxonomy_category_match",
    "reasoning_codes",
    "reasoning_gap",
    "task_supporting_quote",
    "alignment_note",
    "thought_process",
]

HANDLING_HEADERS = [
    "paper_id",
    "strategy_code",
    "pipeline_stage",
    "quantification_metric",
    "handling_supporting_quote",
    "primary_position",
    "handling_gap_identified",
    "internal_consistency",
    "inconsistency_note",
    "thought_process",
]

SUMMARY_HEADERS = [
    "paper_id",
    "model_version",
    "prompt_version",
    "timestamp",
    "subjectivity_def_type",
    "pillar_match",
    "matches_working_definition",
    "definition_gap_identified",
    "task_count",
    "task_names",
    "taxonomy_category_matches",
    "reasoning_codes_all",
    "strategy_codes",
    "primary_position",
    "handling_gap_identified",
    "internal_consistency",
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
ROW_ALIGN = Alignment(vertical="top", wrap_text=True)


def _style_header_row(ws) -> None:
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
    ws.row_dimensions[1].height = 30


def _create_workbook(xlsx_path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, headers in [
        ("Definition", DEFINITION_HEADERS),
        ("Taxonomy", TAXONOMY_HEADERS),
        ("Handling", HANDLING_HEADERS),
        ("Summary", SUMMARY_HEADERS),
    ]:
        ws = wb.create_sheet(sheet_name)
        ws.append(headers)
        _style_header_row(ws)
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 25

    wb.save(xlsx_path)


def _append_row(ws, row: list) -> None:
    ws.append(row)
    for cell in ws[ws.max_row]:
        cell.alignment = ROW_ALIGN


# ---------------------------------------------------------------------------
# Row builders
# ---------------------------------------------------------------------------

def _definition_row(record: VerifierRecord) -> list:
    d = record.definition
    return [
        d.paper_id,
        d.subjectivity_def_type.value,
        ", ".join(p.value for p in d.pillar_match),
        d.matches_working_definition.value,
        d.definition_gap_identified or "",
        d.subjectivity_distinguished_from or "",
        d.supporting_definition_quote or "",
        d.thought_process or "",
    ]


def _taxonomy_rows(record: VerifierRecord) -> list[list]:
    rows = []
    t = record.taxonomy
    if not t.focal_tasks:
        rows.append([
            t.paper_id, "", "", "", "", "", "", "",
            t.thought_process or "",
        ])
        return rows
    for task in t.focal_tasks:
        rows.append([
            t.paper_id,
            task.task_name,
            task.author_task_label.value,
            task.taxonomy_category_match.value,
            ", ".join(r.value for r in task.reasoning_codes),
            task.reasoning_gap or "",
            task.task_supporting_quote or "",
            task.alignment_note or "",
            t.thought_process or "",
        ])
    return rows


def _handling_rows(record: VerifierRecord) -> list[list]:
    rows = []
    h = record.handling
    if not h.handling_strategies:
        rows.append([
            h.paper_id, "", "", "", "",
            h.primary_position.value,
            h.handling_gap_identified or "",
            h.internal_consistency.value,
            h.inconsistency_note or "",
            h.thought_process or "",
        ])
        return rows
    for strategy in h.handling_strategies:
        rows.append([
            h.paper_id,
            strategy.strategy_code.value,
            strategy.pipeline_stage.value,
            strategy.quantification_metric or "",
            strategy.handling_supporting_quote or "",
            h.primary_position.value,
            h.handling_gap_identified or "",
            h.internal_consistency.value,
            h.inconsistency_note or "",
            h.thought_process or "",
        ])
    return rows


def _summary_row(record: VerifierRecord) -> list:
    d = record.definition
    t = record.taxonomy
    h = record.handling

    task_names = "; ".join(task.task_name for task in t.focal_tasks)
    taxonomy_matches = "; ".join(task.taxonomy_category_match.value for task in t.focal_tasks)
    all_reasoning = "; ".join(
        ", ".join(r.value for r in task.reasoning_codes)
        for task in t.focal_tasks
    )
    strategy_codes = ", ".join(s.strategy_code.value for s in h.handling_strategies)

    return [
        record.paper_id,
        record.model_version,
        record.prompt_version,
        record.timestamp,
        d.subjectivity_def_type.value,
        ", ".join(p.value for p in d.pillar_match),
        d.matches_working_definition.value,
        d.definition_gap_identified or "",
        len(t.focal_tasks),
        task_names,
        taxonomy_matches,
        all_reasoning,
        strategy_codes,
        h.primary_position.value,
        h.handling_gap_identified or "",
        h.internal_consistency.value,
    ]


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def load_processed_ids(xlsx_path: Path) -> set[str]:
    if not xlsx_path.exists():
        return set()
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb["Summary"]
    ids = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            ids.add(str(row[0]))
    wb.close()
    return ids


def append_to_xlsx(record: VerifierRecord, xlsx_path: Path) -> None:
    if not xlsx_path.exists():
        _create_workbook(xlsx_path)

    wb = load_workbook(xlsx_path)

    _append_row(wb["Definition"], _definition_row(record))

    for row in _taxonomy_rows(record):
        _append_row(wb["Taxonomy"], row)

    for row in _handling_rows(record):
        _append_row(wb["Handling"], row)

    _append_row(wb["Summary"], _summary_row(record))

    wb.save(xlsx_path)